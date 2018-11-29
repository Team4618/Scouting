import json
from os import fsencode, fsdecode, listdir, makedirs
from os.path import isfile, isdir
from struct import pack as stpack
from threading import Thread
from tkinter import *
from tkinter.ttk import *

import bluetooth
import openpyxl as excel
from pyqrcode import create as createQR

import GUI

# static vars
uuid = "cb3bd26c-4436-11e8-842f-0ed5f89f718b"

verification = "4618 SCOUTING APP"

textOutput = None  # this is used to output text. None shouldn't be a problem because its declared in GUI init


class ScoutingUI:
    def __init__(self, parent, *args):
        self.page = Frame(parent)
        parent.add(self.page, text="Scouting")

        # all of our output that would normally go on console will go here
        outputScrollBar = Scrollbar(self.page)
        outputScrollBar.grid(row=0, column=1, rowspan=999, sticky=NS)

        self.output = Text(self.page, background="black", foreground="white", yscrollcommand=outputScrollBar.set)
        self.output.grid(row=0, column=0, rowspan=999)
        outputScrollBar.config(command=self.output.yview)

        global textOutput
        textOutput = self.output

        #####

        try:
            adrr = bluetooth.read_local_bdaddr()[0]
        except OSError:
            adrr = "No bluetooth address found. Maybe bluetooth isn't turned on?"

        # create qr code from address
        self.qr = createQR(adrr)

        # setup tkinter image with the qr code
        self.qrBm = BitmapImage(data=self.qr.xbm(scale=5))
        self.qrBm.config(background="white")
        Label(self.page, image=self.qrBm).grid(row=2, column=2)
        Label(self.page, text=adrr).grid(row=3, column=2)

        #####

        # start button
        global clicked
        clicked = False

        self.startButtonText = StringVar()
        self.startButtonText.set("Start")

        self.startButton = Button(self.page, textvariable=self.startButtonText, command=self.startBtnClick)

        self.startButton.grid(row=1, column=2, columnspan=2, sticky=EW)

        #####

        # verification textbox label
        Label(self.page, text="Verification:").grid(row=4, column=2)

        # textbox to set verification

        self.verificationText = StringVar()
        self.verificationText.set(verification)
        self.verificationText.trace("w", self.setVerification)
        verificationTextBox = Entry(self.page, textvariable=self.verificationText)
        verificationTextBox.grid(row=5, column=2)

        # generate excel button
        generateExcelButton = Button(self.page, text="Generate Excel", command=generateExcel)
        generateExcelButton.grid(row=6, column=2)

    def setVerification(self, *args):
        global verification
        verification = self.verificationText.get()

    def startBtnClick(self):
        global clicked
        if clicked:
            printtoGUI("Stopping...")
            socket.close()
            printtoGUI("Stopped")
            clicked = False

            printtoGUI()
            self.startButtonText.set("Start")
        else:
            prepExcel()

            Thread(target=startBT).start()
            clicked = True
            self.startButtonText.set("Stop")


def printtoGUI(obj=""):
    textOutput.insert(END, str(obj) + '\n')


def startBT():
    global socket

    printtoGUI("Creating BT socket")

    socket = bluetooth.BluetoothSocket(bluetooth.RFCOMM)
    socket.bind(('', bluetooth.PORT_ANY))
    socket.listen(10)  # will accept max 10 connections at a time

    # tell other devices about our service, providing UUID and specifying a serial connection
    bluetooth.advertise_service(socket, "4618 Scouting Server", service_id=uuid, service_classes=[uuid, bluetooth.
                                SERIAL_PORT_CLASS], profiles=[bluetooth.SERIAL_PORT_PROFILE])

    while 1:
        printtoGUI("Waiting for connection...")
        try:
            conn = socket.accept()[0]
        except OSError:  # usually the socket was closed from somewhere else
            socket.close()
            return

        printtoGUI("Connection received")
        Thread(target=handleConnection, args=(conn,)).start()

    socket.close()


def handleConnection(s):
    # first read: get verification
    # is it necessary? idk
    while 1:
        # read data
        data = s.recv(1024)

        if not data:
            # data is null
            printtoGUI("Connection closed remotely")
            s.close()
            return

        if data.decode("utf-8") != verification:
            printtoGUI("Verification unsuccessful, closing")
            s.close()
            return

        printtoGUI("Verification successful")
        s.send(verification)  # send verification back
        break

    # load template to send later
    with open("template.json", "r") as f:
        template = json.load(f)  # load as json to get rid of whitespace and make message smaller

    templateStr = json.dumps(template)

    sucess = False
    while not sucess:  # loop until client recives template
        printtoGUI("sending template")
        # send template with questions in it
        s.send(stpack("!i", len(templateStr)))  # send length of message

        s.send(templateStr)

        '''# read 1.5: get verification
        while 1:
            # read data
            data = s.recv(1)

            if not data:
                # data is null
                printtoGUI("Connection closed remotely")
                s.close()
                return

            if data == 'Y':
                print(data)
                sucess = True
                break
            elif data == 'N':
                sucess = False
                break'''
        sucess = True

    printtoGUI("sent")
    # second read: get the data
    lastRead = ''
    lastReadEquals = False
    while 1:
        # read data
        data = s.recv(1024).decode("utf-8")

        if data == lastRead:
            if lastReadEquals:
                # safe to assume the connection was closed
                printtoGUI("Connection closed remotely")
                s.close()
                return

            lastReadEquals = True
            lastRead = data
            # no need to process identical data
            continue

        lastRead = data

        if not data:
            # probably all null, ignore it
            continue

        if data == "":
            # we can't do anything with an empty string, and it'll make the JSON parser angry later
            continue

        try:
            msgJSON = json.loads(data)
        except ValueError:
            printtoGUI("Message not JSON formatted: " + data)
            continue

        match = str(msgJSON["match"])

        # check if the json file for that match already exists, and if it doesn't, make it
        jArray = [msgJSON]
        if isfile(GUI.filedir + "/" + match + ".json"):
            with open(GUI.filedir + "/" + match + ".json", "r") as f:
                tmp = json.load(f)

                if isinstance(jArray, list):
                    jArray = tmp
                    jArray.append(msgJSON)

        with open(GUI.filedir + "/" + match + ".json", "w") as f:
            json.dump(jArray, f, indent=4)

        printtoGUI("Wrote data to " + match + ".json")

        updateExcel(msgJSON)


def prepExcel():
    global wb
    wb = excel.Workbook()

    if isfile(GUI.filedir + "/output/teamsSummary.xlsx"):
        wb = excel.load_workbook(GUI.filedir + "/output/teamsSummary.xlsx")
    if not isdir(GUI.filedir + "/output/"):
        makedirs(GUI.filedir + "/output/")


def updateExcel(data, save=True):
    # update the excel file
    team = str(data['robot'])
    if team in wb.sheetnames:
        ws = wb[team]
    else:
        ws = wb.create_sheet(team)

        # setup first row
        i = 1
        for key in data:
            cell = ws.cell(row=1, column=i)
            cell.value = key
            i += 1

    # we're gonna write to row match# + 1 because first row is labels
    i = 1
    for key in data:
        cell = ws.cell(row=data['match'] + 1, column=i)
        cell.value = data[key]
        i += 1

    if save:
        wb.save(GUI.filedir + "/output/teamsSummary.xlsx")


def generateExcel():
    prepExcel()

    # loop through json files and make an excel doc
    for file in listdir(fsencode(GUI.filedir)):
        fileName = fsdecode(file)

        if fileName.lower().endswith(".json"):
            with open(GUI.filedir + '/' + fileName) as f:
                fileJson = json.load(f)

                for i in fileJson:
                    try:
                        updateExcel(i, False)

                    except TypeError:  # whatever we just encountered wasn't a dict, so we should ignore it
                        continue
                    except KeyError:  # this was a dict, but not one which contains our JSON data
                        continue

    wb.save(GUI.filedir + "/output/teamsSummary.xlsx")
